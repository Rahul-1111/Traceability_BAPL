from django.shortcuts import render
from django.http import JsonResponse
from .models import TraceabilityData
import logging
from .qr_utils import generate_qr_code
import random
import datetime
from pymodbus.client import ModbusTcpClient
import time
import threading
from .plc_utils import PLC_MAPPING
from django.utils.timezone import localtime  # ✅ Added

logger = logging.getLogger(__name__)

# ✅ Shared dictionary for storing PLC statuses
plc_statuses = {}

def monitor_plcs():
    """Continuously monitor PLC connections in a background thread."""
    global plc_statuses
    while True:
        temp_statuses = {}
        for station, plc in PLC_MAPPING.items():
            plc_ip = plc.get("ip")
            if not plc_ip:
                continue

            client = ModbusTcpClient(plc_ip, port=5007, timeout=1)
            is_connected = client.connect()
            client.close()

            temp_statuses[station] = "connected" if is_connected else "disconnected"

        plc_statuses = temp_statuses
        time.sleep(1)

# ✅ Start monitoring thread once
plc_monitor_thread = threading.Thread(target=monitor_plcs, daemon=True)
plc_monitor_thread.start()


def plc_status(request):
    """Return PLC statuses with combined stations for shared PLCs."""
    combined_statuses = {
        "St 1 & 2": plc_statuses.get("st1", "disconnected"),
        "St 3":     plc_statuses.get("st3", "disconnected"),
        "St 4":     plc_statuses.get("st4", "disconnected"),
        "St 5 & 6": plc_statuses.get("st6", "disconnected"),
    }

    connected_count    = sum(1 for s in combined_statuses.values() if s == "connected")
    disconnected_count = sum(1 for s in combined_statuses.values() if s == "disconnected")

    return JsonResponse({
        "plc_statuses":       combined_statuses,
        "connected_count":    connected_count,
        "disconnected_count": disconnected_count
    })


# ✅ Render main page
def combined_page(request):
    return render(request, 'track/combined_page.html')


# ✅ QR Code Generation View
def generate_qr_code_view(request):
    if request.method == "POST":
        prefix = request.POST.get("prefix")
        serial_number = random.randint(10000, 99999)

        if not prefix:
            return JsonResponse({"error": "Prefix is required"}, status=400)

        try:
            response_message = generate_qr_code(prefix, serial_number)
            return JsonResponse({
                "message": response_message,
                "generated_code": f"{prefix}-{datetime.datetime.now().strftime('%d%m%y')}-{serial_number}"
            })
        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            return JsonResponse({"error": str(e)}, status=500)


# ✅ Fetch Torque Data
from datetime import date
from django.db.models import Q

def fetch_torque_data(request):
    if request.method == "GET":
        today = date.today()

        # ✅ Today's records
        today_records = TraceabilityData.objects.filter(date=today)

        # ✅ NOT OK or NULL/blank records
        not_ok_or_null_records = TraceabilityData.objects.filter(
            Q(st1_result="NOT OK") | Q(st2_result="NOT OK") |
            Q(st3_result="NOT OK") | Q(st4_result="NOT OK") |
            Q(st5_result="NOT OK") | Q(st6_result="NOT OK") |
            Q(st1_result__isnull=True) | Q(st2_result__isnull=True) |
            Q(st3_result__isnull=True) | Q(st4_result__isnull=True) |
            Q(st5_result__isnull=True) | Q(st6_result__isnull=True) |
            Q(st1_result="") | Q(st2_result="") | Q(st3_result="") |
            Q(st4_result="") | Q(st5_result="") | Q(st6_result="")
        )

        # ✅ Combine & order newest first
        combined_data = (today_records | not_ok_or_null_records).distinct().order_by('-date', '-time')

        formatted_data = [
            {
                "part_number": item.part_number,
                "date":        item.date.strftime("%Y-%m-%d") if item.date else "",
                "time":        item.time.strftime("%H:%M:%S") if item.time else "",
                "shift":       item.shift,

                # ✅ Station results
                "st1_result":  item.st1_result,
                "st2_result":  item.st2_result,
                "st3_result":  item.st3_result,
                "st4_result":  item.st4_result,
                "st5_result":  item.st5_result,
                "st6_result":  item.st6_result,

                # ✅ Station datetimes — localtime() converts UTC → IST
                "st1_datetime": localtime(item.st1_datetime).strftime("%Y-%m-%d %H:%M:%S") if item.st1_datetime else "",
                "st2_datetime": localtime(item.st2_datetime).strftime("%Y-%m-%d %H:%M:%S") if item.st2_datetime else "",
                "st3_datetime": localtime(item.st3_datetime).strftime("%Y-%m-%d %H:%M:%S") if item.st3_datetime else "",
                "st4_datetime": localtime(item.st4_datetime).strftime("%Y-%m-%d %H:%M:%S") if item.st4_datetime else "",
                "st5_datetime": localtime(item.st5_datetime).strftime("%Y-%m-%d %H:%M:%S") if item.st5_datetime else "",
                "st6_datetime": localtime(item.st6_datetime).strftime("%Y-%m-%d %H:%M:%S") if item.st6_datetime else "",
            }
            for item in combined_data
        ]

        return JsonResponse({"data": formatted_data})


# ✅ Search Parts View
from .filters import TraceabilityDataFilter

def search_parts(request):
    queryset = TraceabilityData.objects.all()
    f = TraceabilityDataFilter(request.GET, queryset=queryset)
    return render(request, 'track/search_parts.html', {'filter': f})


# ✅ Export to Excel — IST datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def export_to_excel(request):
    queryset = TraceabilityData.objects.all()
    f = TraceabilityDataFilter(request.GET, queryset=queryset)
    filtered_data = f.qs

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Traceability Data"

    headers = [
        "SR No", "Part Number", "Date", "Time", "Shift",
        "DRILLING ST1", "ST1 DATETIME",
        "VACUUM ST2",   "ST2 DATETIME",
        "HOT PLATE ST3","ST3 DATETIME",
        "LEAKAGE ST4",  "ST4 DATETIME",
        "POKA YOKE ST5","ST5 DATETIME",
        "WEIGHT ST6",   "ST6 DATETIME",
    ]
    ws.append(headers)

    for obj in filtered_data:
        ws.append([
            obj.sr_no,
            obj.part_number,
            obj.date.strftime("%Y-%m-%d") if obj.date else "",
            obj.time.strftime("%H:%M:%S") if obj.time else "",
            obj.shift,

            # ✅ ST1 — localtime for Excel too
            obj.st1_result,
            localtime(obj.st1_datetime).strftime("%Y-%m-%d %H:%M:%S") if obj.st1_datetime else "",

            # ✅ ST2
            obj.st2_result,
            localtime(obj.st2_datetime).strftime("%Y-%m-%d %H:%M:%S") if obj.st2_datetime else "",

            # ✅ ST3
            obj.st3_result,
            localtime(obj.st3_datetime).strftime("%Y-%m-%d %H:%M:%S") if obj.st3_datetime else "",

            # ✅ ST4
            obj.st4_result,
            localtime(obj.st4_datetime).strftime("%Y-%m-%d %H:%M:%S") if obj.st4_datetime else "",

            # ✅ ST5
            obj.st5_result,
            localtime(obj.st5_datetime).strftime("%Y-%m-%d %H:%M:%S") if obj.st5_datetime else "",

            # ✅ ST6
            obj.st6_result,
            localtime(obj.st6_datetime).strftime("%Y-%m-%d %H:%M:%S") if obj.st6_datetime else "",
        ])

    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="traceability_data.xlsx"'
    wb.save(response)
    return response